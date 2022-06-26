from django.shortcuts import render
from dataclasses import dataclass
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.template import loader
import openpyxl
from .static.webapp1 import *


# Create your views here.

class Datawiki:
	id:str = None
	name:str = None
	description:str = None
	link:str = None
	category:str = None
	field:str = None
	task:str = None
	instances:str = None
	num:str = None
	tutorial:str = None
	index:int = None

ex= Datawiki()
arr = []
for i in range(0, 60):
	ex = Datawiki()
	arr.append(ex)

def make():
	wb = openpyxl.load_workbook('C:\\Users\\user\\WebProject-django\\webapp1\\static\\webapp1\\datawikiinfo.xlsx')
	sheet = wb.get_sheet_by_name('Sheet1')
	for i in range (2,62):  # 2~61
		for j in range (1, 11):  # A~J
			arr[i-2].index = i-2  # 0~59 인덱스
			if (j==1):
				arr[i-2].id = sheet.cell(row=i, column=j).value
			elif (j==2):
				arr[i-2].name = sheet.cell(row=i, column=j).value
			elif (j==3):
				arr[i-2].description = sheet.cell(row=i, column=j).value
			elif (j==4):
				arr[i-2].link = sheet.cell(row=i, column=j).value
			elif (j==5):
				arr[i-2].category = sheet.cell(row=i, column=j).value
			elif (j==6):
				arr[i-2].field = sheet.cell(row=i, column=j).value
			elif (j==7):
			    arr[i-2].task = sheet.cell(row=i, column=j).value
			elif (j==8):
				arr[i-2].instances = sheet.cell(row=i, column=j).value
			elif (j==9):
				arr[i-2].num = sheet.cell(row=i, column=j).value
			else:
				arr[i-2].tutorial = sheet.cell(row=i, column=j).value

def goData(request):
	make()
	template = loader.get_template('webapp1/DataWiki.html')
	context = {
		'data_list': arr  # 딕셔너리 키:값. 이는 파이썬 코드인데 html 코드로 전달해주기 위해 나중에 템플릿 안의 html 파일에서 {{ corrent_date }} 이렇게 사용하여
                    	  # 장고 템플릿 태그로 {{ }} 로 중괄호 2개를 감싸서 사용한다.
    }
	return HttpResponse(template.render(context, request))

def htmlgoHome(request):
	template = loader.get_template('webapp1/Home.html')
	context = {
	}
	return HttpResponse(template.render(context, request))

def htmlgoForum(request):
	template = loader.get_template('webapp1/Forum.html')
	context = {
	}
	return HttpResponse(template.render(context, request))

def htmlgoShareHub(request):
	template = loader.get_template('webapp1/ShareHub.html')
	context = {
	}
	return HttpResponse(template.render(context, request))